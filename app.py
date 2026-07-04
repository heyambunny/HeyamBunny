"""
Attendance Analyzer Application
Complete working version - Final
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, time, timedelta
from io import BytesIO
from typing import Tuple, List, Optional, Dict, Any
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(
    page_title="Attendance Analyzer",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# DATA PROCESSING
# ============================================================================

class DataProcessor:
    """Handles data loading, cleaning, and validation."""
    
    @staticmethod
    def parse_date_time_columns(df: pd.DataFrame) -> pd.DataFrame:
        """Parse date and time columns from the attendance file."""
        df_clean = df.copy()
        
        # Parse date columns - FIXED: dayfirst=True for dd-mm-yyyy format
        date_columns = ['Attendance Date', 'In Date', 'Out Date']
        for col in date_columns:
            if col in df_clean.columns:
                df_clean[col] = pd.to_datetime(df_clean[col], format='mixed', dayfirst=True, errors='coerce')
        
        # Parse time columns
        time_columns = ['In Time', 'Out Time', 'Shift Start Time', 'Shift End Time']
        for col in time_columns:
            if col in df_clean.columns:
                try:
                    if pd.api.types.is_object_dtype(df_clean[col]):
                        df_clean[col] = pd.to_datetime(df_clean[col], format='mixed', errors='coerce').dt.time
                    elif pd.api.types.is_datetime64_any_dtype(df_clean[col]):
                        df_clean[col] = df_clean[col].dt.time
                except:
                    df_clean[col] = None
        
        return df_clean
    
    @staticmethod
    def clean_data(df: pd.DataFrame) -> pd.DataFrame:
        """Clean and prepare data."""
        df_clean = df.dropna(subset=['Employee Code', 'Attendance Date'])
        df_clean['Employee Code'] = df_clean['Employee Code'].astype(str).str.strip()
        df_clean = df_clean.drop_duplicates(subset=['Employee Code', 'Attendance Date', 'In Time', 'Out Time'])
        df_clean = df_clean.sort_values(['Employee Code', 'Attendance Date']).reset_index(drop=True)
        return df_clean
    
    @staticmethod
    def validate_columns(df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """Check if required columns exist."""
        required = ['Employee Code', 'Employee Name', 'Department', 'Designation', 
                   'Attendance Date', 'In Time', 'Out Time']
        missing = [col for col in required if col not in df.columns]
        return len(missing) == 0, missing
    
    @staticmethod
    def get_date_range(df: pd.DataFrame) -> Tuple[datetime, datetime]:
        """Get min and max dates from the file."""
        return df['Attendance Date'].min(), df['Attendance Date'].max()

# ============================================================================
# ATTENDANCE ANALYZER
# ============================================================================

class AttendanceAnalyzer:
    """Main attendance analysis logic."""
    
    WEEKDAY_MAP = {
        'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3,
        'Friday': 4, 'Saturday': 5, 'Sunday': 6
    }
    
    def __init__(self, df: pd.DataFrame, config: Dict[str, Any]):
        self.df = df
        self.config = config
    
    def is_senior(self, designation: str) -> bool:
        """Check if employee is senior."""
        return designation in self.config['senior_designations']
    
    def get_expected_times(self, is_senior: bool) -> Tuple[time, time]:
        """Get expected in/out times based on seniority."""
        if is_senior:
            return self.config['senior_first_half'], self.config['senior_second_half']
        return self.config['normal_first_half'], self.config['normal_second_half']
    
    def get_employee_attendance(self, emp_code: str, date: datetime) -> Optional[Dict]:
        """Get attendance for specific employee on specific date."""
        target_date = pd.Timestamp(date).normalize()
        
        records = self.df[
            (self.df['Employee Code'] == emp_code) & 
            (self.df['Attendance Date'].dt.normalize() == target_date)
        ]
        
        if records.empty:
            return None
        
        # Multiple punches: earliest In, latest Out
        in_times = records['In Time'].dropna()
        out_times = records['Out Time'].dropna()
        
        if in_times.empty and out_times.empty:
            return None
        
        return {
            'in_time': min(in_times) if not in_times.empty else None,
            'out_time': max(out_times) if not out_times.empty else None,
            'in_date': records['In Date'].iloc[0] if 'In Date' in records.columns else None,
            'out_date': records['Out Date'].iloc[0] if 'Out Date' in records.columns else None,
            'working_hours': records['Working Hours'].iloc[0] if 'Working Hours' in records.columns else None,
            'shift': records['Shift'].iloc[0] if 'Shift' in records.columns else None
        }

    def check_compliance(self, in_time: Optional[time], out_time: Optional[time], 
                        is_senior: bool) -> Tuple[bool, bool, str]:
        """Check first half and second half compliance."""
        
        expected_in, expected_out = self.get_expected_times(is_senior)
        
        # Helper to convert anything to minutes
        def time_to_minutes(t):
            if t is None:
                return None
            if isinstance(t, str):
                try:
                    parts = str(t).split(':')
                    return int(parts[0]) * 60 + int(parts[1])
                except:
                    return None
            try:
                return t.hour * 60 + t.minute
            except:
                return None
        
        in_mins = time_to_minutes(in_time)
        out_mins = time_to_minutes(out_time)
        expected_in_mins = time_to_minutes(expected_in)
        expected_out_mins = time_to_minutes(expected_out)
        
        first_half = True if in_mins is None else in_mins > expected_in_mins
        second_half = True if out_mins is None else out_mins < expected_out_mins
        
        if first_half and second_half:
            status = "Full Day Issue"
        elif first_half:
            status = "First Half Issue"
        elif second_half:
            status = "Second Half Issue"
        else:
            status = "Present"
        
        return first_half, second_half, status
    
    def get_employee_details(self, emp_code: str, departments: List[str], 
                            weekdays: List[str]) -> pd.DataFrame:
        """Get detailed date-wise attendance for one employee."""
        
        all_data = self.df[self.df['Department'].isin(departments)]
        all_dates = all_data['Attendance Date'].dropna().unique()
        
        weekday_nums = [self.WEEKDAY_MAP[w] for w in weekdays]
        relevant_dates = sorted([d for d in all_dates if d.weekday() in weekday_nums])
        
        emp_data = self.df[self.df['Employee Code'] == emp_code]
        if emp_data.empty:
            return pd.DataFrame()
        
        emp_info = emp_data.iloc[0]
        is_senior = self.is_senior(emp_info.get('Designation', ''))
        expected_in, expected_out = self.get_expected_times(is_senior)
        
        # Safe formatter
        def safe_str(val):
            if val is None:
                return '-'
            try:
                return val.strftime('%H:%M')
            except:
                return str(val)
        
        def safe_date_str(val, fmt):
            try:
                return val.strftime(fmt)
            except:
                return str(val)[:10]
        
        detail_rows = []
        
        for date in relevant_dates:
            attendance = self.get_employee_attendance(emp_code, date)
            
            if attendance is None:
                detail_rows.append({
                    'Date': safe_date_str(date, '%d-%b-%Y'),
                    'Day': safe_date_str(date, '%A'),
                    'Status': '❌ ABSENT',
                    'Punch In': '-',
                    'Punch Out': '-',
                    'Expected In': safe_str(expected_in),
                    'Expected Out': safe_str(expected_out),
                    'Issue Type': 'Leave - No Record',
                    'Working Hours': '-',
                    'Shift': '-'
                })
            else:
                fh, sh, status = self.check_compliance(
                    attendance['in_time'], 
                    attendance['out_time'], 
                    is_senior
                )
                
                issue_type = []
                if fh:
                    issue_type.append('First Half (Late)')
                if sh:
                    issue_type.append('Second Half (Early)')
                
                detail_rows.append({
                    'Date': safe_date_str(date, '%d-%b-%Y'),
                    'Day': safe_date_str(date, '%A'),
                    'Status': f'✅ Present' if status == 'Present' else f'⚠️ {status}',
                    'Punch In': safe_str(attendance['in_time']),
                    'Punch Out': safe_str(attendance['out_time']),
                    'Expected In': safe_str(expected_in),
                    'Expected Out': safe_str(expected_out),
                    'Issue Type': ', '.join(issue_type) if issue_type else 'None',
                    'Working Hours': attendance.get('working_hours', '-'),
                    'Shift': attendance.get('shift', '-')
                })
        
        return pd.DataFrame(detail_rows)
    
    def generate_summary_report(self, departments: List[str], employees: List[str], 
                               weekdays: List[str]) -> pd.DataFrame:
        """Generate summary report with counts."""
        
        mask = (self.df['Department'].isin(departments)) & (self.df['Employee Code'].isin(employees))
        filtered_df = self.df[mask].copy()
        
        if filtered_df.empty:
            return pd.DataFrame()
        
        # Get ALL dates from the file
        all_dates = filtered_df['Attendance Date'].dropna().unique()
        weekday_nums = [self.WEEKDAY_MAP[w] for w in weekdays]
        relevant_dates = sorted([d for d in all_dates if d.weekday() in weekday_nums])
        
        unique_employees = filtered_df['Employee Code'].unique()
        
        report_rows = []
        progress_bar = st.progress(0)
        
        for idx, emp_code in enumerate(unique_employees):
            progress_bar.progress((idx + 1) / len(unique_employees))
            
            emp_info = filtered_df[filtered_df['Employee Code'] == emp_code].iloc[0]
            
            row = {
                'Employee Code': emp_code,
                'Employee Name': emp_info.get('Employee Name', ''),
                'Department': emp_info.get('Department', ''),
                'Designation': emp_info.get('Designation', '')
            }
            
            total_issues = 0
            is_senior = self.is_senior(emp_info.get('Designation', ''))
            
            for weekday in weekdays:
                weekday_num = self.WEEKDAY_MAP[weekday]
                weekday_dates = [d for d in relevant_dates if d.weekday() == weekday_num]
                
                leave = 0
                first_half = 0
                second_half = 0
                
                for date in weekday_dates:
                    attendance = self.get_employee_attendance(emp_code, date)
                    
                    if attendance is None:
                        leave += 1
                    else:
                        fh, sh, _ = self.check_compliance(
                            attendance['in_time'], 
                            attendance['out_time'], 
                            is_senior
                        )
                        if fh:
                            first_half += 1
                        if sh:
                            second_half += 1
                
                row[f'{weekday} Leave'] = leave
                row[f'{weekday} First Half'] = first_half
                row[f'{weekday} Second Half'] = second_half
                
                total_issues += leave + first_half + second_half
            
            row['Total Issues'] = total_issues
            report_rows.append(row)
        
        progress_bar.empty()
        
        report_df = pd.DataFrame(report_rows)
        if not report_df.empty:
            report_df = report_df.sort_values('Total Issues', ascending=False).reset_index(drop=True)
        
        return report_df

# ============================================================================
# EXCEL EXPORT
# ============================================================================

class ExcelExporter:
    """Export report to formatted Excel."""
    
    @staticmethod
    def export_summary(df: pd.DataFrame) -> bytes:
        """Export summary DataFrame to Excel bytes."""
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Summary', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Summary']
            
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for column in worksheet.columns:
                max_length = 0
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
        
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def export_detailed(detail_dict: Dict[str, pd.DataFrame], df_original: pd.DataFrame) -> bytes:
        """Export detailed reports with multiple sheets."""
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for emp_code, df_detail in detail_dict.items():
                emp_info = df_original[df_original['Employee Code'] == emp_code]
                emp_name = emp_info['Employee Name'].iloc[0] if not emp_info.empty else ''
                emp_dept = emp_info['Department'].iloc[0] if not emp_info.empty else ''
                emp_desig = emp_info['Designation'].iloc[0] if not emp_info.empty else ''
                
                sheet_name = str(emp_code)[:31]
                
                # Write employee info at top
                from openpyxl.styles import Font, PatternFill, Alignment
                
                info_data = [
                    ['Employee Code:', emp_code],
                    ['Employee Name:', emp_name],
                    ['Department:', emp_dept],
                    ['Designation:', emp_desig],
                    ['', ''],
                ]
                
                for i, row_data in enumerate(info_data, start=1):
                    worksheet_temp = writer.sheets.get(sheet_name)
                
                # Create sheet with employee info first
                info_df = pd.DataFrame(info_data)
                info_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=0)
                
                # Write detail data below
                df_detail.to_excel(writer, sheet_name=sheet_name, index=False, startrow=6)
                
                worksheet = writer.sheets[sheet_name]
                
                # Format employee info
                bold_font = Font(bold=True, size=12)
                for row in range(1, 5):
                    worksheet.cell(row=row, column=1).font = bold_font
                    worksheet.cell(row=row, column=2).font = Font(size=12)
                
                # Format detail headers (row 7)
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True, size=11)
                
                for cell in worksheet[7]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-fit columns
                for column in worksheet.columns:
                    max_length = 0
                    for cell in column:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
        
        output.seek(0)
        return output.getvalue()

# ============================================================================
# STREAMLIT UI
# ============================================================================

def main():
    """Main Streamlit application."""
    
    st.title("📊 Attendance Analyzer")
    st.markdown("---")
    
    # Initialize session state
    if 'df' not in st.session_state:
        st.session_state.df = None
    if 'summary_report' not in st.session_state:
        st.session_state.summary_report = None
    if 'detailed_reports' not in st.session_state:
        st.session_state.detailed_reports = None
    
    # ===== STEP 1: UPLOAD =====
    st.header("1. Upload Attendance File")
    
    uploaded_file = st.file_uploader(
        "Choose Excel file",
        type=['xlsx', 'xls'],
        help="Upload attendance Excel file with columns: Employee Code, Employee Name, Department, Designation, Attendance Date, In Time, Out Time"
    )
    
    if uploaded_file:
        try:
            with st.spinner("Processing file..."):
                df_raw = pd.read_excel(uploaded_file)
                
                processor = DataProcessor()
                is_valid, missing = processor.validate_columns(df_raw)
                
                if not is_valid:
                    st.error(f"❌ Missing required columns: {', '.join(missing)}")
                    st.info(f"📋 Columns found: {', '.join(df_raw.columns.tolist())}")
                else:
                    df_parsed = processor.parse_date_time_columns(df_raw)
                    df_clean = processor.clean_data(df_parsed)
                    
                    st.session_state.df = df_clean
                    
                    start, end = processor.get_date_range(df_clean)
                    unique_dates = df_clean['Attendance Date'].dropna().nunique()
                    
                    st.success(f"""
                    ✅ File loaded successfully!
                    - **Records**: {len(df_clean):,}
                    - **Employees**: {df_clean['Employee Code'].nunique():,}
                    - **Date Range**: {start.strftime('%d-%b-%Y')} to {end.strftime('%d-%b-%Y')}
                    - **Unique Dates**: {unique_dates}
                    - **Departments**: {df_clean['Department'].nunique()}
                    """)
        except Exception as e:
            st.error(f"❌ Error processing file: {e}")
    
    if st.session_state.df is None:
        st.stop()
    
    df = st.session_state.df
    
    # ===== STEP 2: CONFIGURATION =====
    st.header("2. Configure Timings")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Normal Employees")
        normal_in = st.time_input("First Half - In By", value=time(9, 30), 
                                  help="Employee must punch IN at or before this time")
        normal_out = st.time_input("Second Half - Out After", value=time(18, 0),
                                   help="Employee must punch OUT at or after this time")
    
    with col2:
        st.subheader("Senior Employees")
        senior_in = st.time_input("First Half - In By", value=time(10, 0), key='sin',
                                  help="Senior must punch IN at or before this time")
        senior_out = st.time_input("Second Half - Out After", value=time(17, 30), key='sout',
                                   help="Senior must punch OUT at or after this time")
    
    st.markdown("---")
    st.subheader("Senior Designations")
    all_designations = sorted(df['Designation'].dropna().unique().tolist())
    senior_designations = st.multiselect(
        "Select which designations are Senior",
        options=all_designations,
        help="These employees will use Senior timings instead of Normal timings"
    )
    
    # ===== STEP 3: FILTERS =====
    st.header("3. Filters")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.subheader("Department")
        all_depts = sorted(df['Department'].dropna().unique().tolist())
        selected_depts = st.multiselect(
            "Select Departments",
            options=all_depts,
            default=all_depts
        )
    
    with col2:
        st.subheader("Employee")
        if selected_depts:
            emp_df = df[df['Department'].isin(selected_depts)]
        else:
            emp_df = df
        
        all_emps = sorted(emp_df['Employee Code'].unique().tolist())
        
        emp_labels = []
        emp_map = {}
        for code in all_emps:
            name = emp_df[emp_df['Employee Code'] == code]['Employee Name'].iloc[0]
            label = f"{code} - {name}"
            emp_labels.append(label)
            emp_map[label] = code
        
        selected_labels = st.multiselect(
            "Select Employees",
            options=emp_labels,
            default=emp_labels
        )
        selected_emps = [emp_map[l] for l in selected_labels]
    
    with col3:
        st.subheader("Weekdays")
        weekdays_list = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        selected_weekdays = st.multiselect(
            "Select Weekdays to Analyze",
            options=weekdays_list,
            default=['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        )
    
    # Show date info
    if selected_weekdays:
        all_dates_in_file = df['Attendance Date'].dropna().unique()
        weekday_nums = [AttendanceAnalyzer.WEEKDAY_MAP[w] for w in selected_weekdays]
        dates_to_check = sorted([d for d in all_dates_in_file if d.weekday() in weekday_nums])
        
        st.info(f"""
        📅 **Analysis Summary:**
        - **Total dates in file**: {len(all_dates_in_file)}
        - **Dates matching selected weekdays**: {len(dates_to_check)}
        - **Selected weekdays**: {', '.join(selected_weekdays)}
        - **Date range**: {df['Attendance Date'].min().strftime('%d-%b-%Y')} to {df['Attendance Date'].max().strftime('%d-%b-%Y')}
        """)
    
    # ===== STEP 4: GENERATE REPORT =====
    st.header("4. Generate Report")
    
    if st.button("🔄 Generate Attendance Report", type="primary", use_container_width=True):
        if not selected_depts:
            st.warning("⚠️ Please select at least one department.")
        elif not selected_emps:
            st.warning("⚠️ Please select at least one employee.")
        elif not selected_weekdays:
            st.warning("⚠️ Please select at least one weekday.")
        else:
            with st.spinner("Analyzing attendance data..."):
                config = {
                    'normal_first_half': normal_in,
                    'normal_second_half': normal_out,
                    'senior_first_half': senior_in,
                    'senior_second_half': senior_out,
                    'senior_designations': senior_designations
                }
                
                analyzer = AttendanceAnalyzer(df, config)
                
                # Generate summary
                summary = analyzer.generate_summary_report(selected_depts, selected_emps, selected_weekdays)
                
                # Generate detailed reports
                detailed = {}
                for emp_code in selected_emps:
                    detail_df = analyzer.get_employee_details(emp_code, selected_depts, selected_weekdays)
                    if not detail_df.empty:
                        detailed[emp_code] = detail_df
                
                if summary.empty:
                    st.warning("⚠️ No data found for selected filters.")
                else:
                    st.session_state.summary_report = summary
                    st.session_state.detailed_reports = detailed
                    st.session_state.selected_weekdays = selected_weekdays
                    st.success(f"✅ Report generated for {len(summary)} employees!")
    
    # ===== DISPLAY SUMMARY REPORT =====
    if st.session_state.summary_report is not None:
        report = st.session_state.summary_report
        selected_weekdays = st.session_state.selected_weekdays
        
        st.markdown("---")
        st.subheader("📈 Summary Report")
        
        # Metrics
        cols = st.columns(4)
        with cols[0]:
            st.metric("Total Employees", len(report))
        with cols[1]:
            total_leaves = sum(report[f'{w} Leave'].sum() for w in selected_weekdays if f'{w} Leave' in report.columns)
            st.metric("Total Leaves", int(total_leaves))
        with cols[2]:
            total_fh = sum(report[f'{w} First Half'].sum() for w in selected_weekdays if f'{w} First Half' in report.columns)
            st.metric("First Half Issues", int(total_fh))
        with cols[3]:
            total_sh = sum(report[f'{w} Second Half'].sum() for w in selected_weekdays if f'{w} Second Half' in report.columns)
            st.metric("Second Half Issues", int(total_sh))
        
        st.markdown("---")
        
        # Summary table
        st.dataframe(
            report.style.map(
                lambda x: 'background-color: #ffcccc; font-weight: bold' if x > 0 else '',
                subset=['Total Issues']
            ),
            use_container_width=True,
            hide_index=True,
            height=400
        )
        
        # Export Summary
        col1, col2 = st.columns(2)
        with col1:
            excel_bytes = ExcelExporter.export_summary(report)
            st.download_button(
                "📥 Download Summary (Excel)",
                excel_bytes,
                f"attendance_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col2:
            if st.session_state.detailed_reports:
                detailed_bytes = ExcelExporter.export_detailed(st.session_state.detailed_reports, df)
                st.download_button(
                    "📥 Download All Details (Excel - Multiple Sheets)",
                    detailed_bytes,
                    f"attendance_detailed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        # ===== DETAILED VIEW =====
        st.markdown("---")
        st.subheader("🔍 Detailed Date-wise Attendance")
        
        if st.session_state.detailed_reports:
            # Employee selector
            emp_options = []
            for code in st.session_state.detailed_reports.keys():
                emp_info = df[df['Employee Code'] == code]
                if not emp_info.empty:
                    name = emp_info['Employee Name'].iloc[0]
                    dept = emp_info['Department'].iloc[0]
                    emp_options.append(f"{code} - {name} ({dept})")
            
            selected_emp_detail = st.selectbox(
                "Select employee to view detailed date-wise attendance",
                options=emp_options
            )
            
            if selected_emp_detail:
                emp_code = selected_emp_detail.split(' - ')[0]
                detail_df = st.session_state.detailed_reports[emp_code]
                
                # Color coding for status
                def color_status(val):
                    if 'ABSENT' in str(val):
                        return 'background-color: #ffcccc; color: #cc0000; font-weight: bold'
                    elif 'Issue' in str(val):
                        return 'background-color: #fff3cd; color: #856404; font-weight: bold'
                    elif 'Present' in str(val):
                        return 'background-color: #d4edda; color: #155724'
                    return ''
                
                styled_detail = detail_df.style.map(color_status, subset=['Status'])
                
                st.dataframe(
                    styled_detail,
                    use_container_width=True,
                    hide_index=True,
                    height=500
                )
                
                # Stats
                total_dates = len(detail_df)
                absent_dates = len(detail_df[detail_df['Status'].str.contains('ABSENT')])
                issue_dates = len(detail_df[detail_df['Status'].str.contains('Issue')])
                present_dates = len(detail_df[detail_df['Status'].str.contains('Present')])
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Dates Checked", total_dates)
                with col2:
                    st.metric("✅ Present", present_dates)
                with col3:
                    st.metric("⚠️ Issues", issue_dates)
                with col4:
                    st.metric("❌ Absent", absent_dates)
                
                # Absent dates list
                if absent_dates > 0:
                    with st.expander(f"📋 Absent Dates List ({absent_dates} days)", expanded=True):
                        absent_df = detail_df[detail_df['Status'].str.contains('ABSENT')][
                            ['Date', 'Day', 'Expected In', 'Expected Out']
                        ]
                        st.dataframe(absent_df, use_container_width=True, hide_index=True)
                
                # Issue details
                if issue_dates > 0:
                    with st.expander(f"⚠️ Issue Details ({issue_dates} days)", expanded=True):
                        issue_df = detail_df[detail_df['Status'].str.contains('Issue')][
                            ['Date', 'Day', 'Punch In', 'Punch Out', 'Expected In', 'Expected Out', 'Issue Type']
                        ]
                        st.dataframe(issue_df, use_container_width=True, hide_index=True)

if __name__ == "__main__":
    main()